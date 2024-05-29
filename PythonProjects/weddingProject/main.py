import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.styles import Font
import tkinter.font as tkFont

import os
import win32api
import win32print
from tkinter import simpledialog

# Function to get the list of available printers
def get_printers():
    printers = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
    return [printer[2] for printer in printers]

# Function to print the Excel file using the selected printer
def print_excel_file(printer_name):
    excel_file_path = '.\data.xlsx'  # Hardcoded file path to the Excel file
    if not os.path.exists(excel_file_path):
        print("The Excel file does not exist at the specified path.")
        return

    # Set the selected printer as default
    win32print.SetDefaultPrinter(printer_name)
    # Print the Excel file using the selected printer
    os.startfile(excel_file_path, "print")


def open_excel():

    # Replace 'path_to_excel_file.xlsx' with the actual path to your Excel file
    os.startfile('.\data.xlsx')
        
# Function to search for the data
def search_data():
    id_search = entry_id.get()
    name_search = entry_name.get()
    load_data(id_search, name_search)  # Pass the ID and Name search terms to load_data


# Function to clear all the entry widgets and reset the Treeview
def clear_entries():
    entry_id.delete(0, tk.END)
    entry_name.delete(0, tk.END)
    entry_usd.delete(0, tk.END)
    entry_riel.delete(0, tk.END)
    load_data()  # Call load_data without arguments to reset the Treeview
    

def update_data():
    global label_prev_id, label_prev_name, label_prev_usd, label_prev_riel
    try:
        id_search = entry_id.get()
        name_to_find = entry_name.get()  # Also get the Name
        usd_value = entry_usd.get()
        riel_value = entry_riel.get()

        # Convert USD and Riel values to integers, defaulting to 0 if empty
        usd_value = int(usd_value) if usd_value else 0
        riel_value = int(riel_value) if riel_value else 0

        # Capture the current data before updating
        prev_id = id_search
        prev_name = name_to_find
        prev_usd = usd_value
        prev_riel = riel_value

        # Update the labels with the previous data
        label_prev_id.config(text=f"Previous ID: {prev_id}")
        label_prev_name.config(text=f"Previous Name: {prev_name}")
        label_prev_usd.config(text=f"Previous USD: {prev_usd}")
        label_prev_riel.config(text=f"Previous Riel: {prev_riel}")

        found = False
        for row in sheet.iter_rows(min_row=2):
            # Check if both ID and Name match before updating
            if str(row[0].value).zfill(3) == id_search.zfill(3) and row[1].value == name_to_find:
                row[2].value = usd_value  # Update the USD value
                row[3].value = riel_value  # Update the Riel value
                found = True
                break  # Break after updating the correct record

        if not found:
            # Add a new row with the new data if no matching record is found
            sheet.append([id_search.zfill(3), name_to_find, usd_value, riel_value])
            print("New data added successfully.")

        # Save the workbook after updating or adding data
        wb.save('.\data.xlsx')
        print("Data updated successfully.")
        refresh_data()
        update_totals()  # Update totals after adding or updating data

    except ValueError:
        print("Please enter a valid number for USD and Riel.")
    except Exception as e:
        print(f"An error occurred: {e}")
# Function to update the totals
def update_totals():
    total_guests = 0
    total_usd = 0
    total_riel = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Convert None to 0 before comparison and addition
        usd = row[2] if row[2] is not None else 0
        riel = row[3] if row[3] is not None else 0

        if usd > 0 or riel > 0:
            total_guests += 1
        total_usd += usd
        total_riel += riel

    label_total_guests.config(text=f"Total Guests: {total_guests}")
    label_total_usd.config(text=f"Total USD: {total_usd}")
    label_total_riel.config(text=f"Total Riel: {total_riel}")
# Function to fill the entry widgets with the data from the selected row
def on_tree_select(event):
    # Check if there is a selection
    if tree.selection():
        selected_item = tree.selection()[0]  # Get selected item
        # Get the values of the selected row
        selected_row = tree.item(selected_item, 'values')
        # Clear the entry fields
        clear_entries()
        # Insert the data into the entry fields
        entry_id.insert(0, selected_row[0])
        entry_name.insert(0, selected_row[1])
        entry_usd.insert(0, selected_row[2])
        entry_riel.insert(0, selected_row[3])
    else:
        print("Item selected")
# Function to load data from Excel into the Treeview with search functionality
def load_data(id_search=None, name_search=None):
    for item in tree.get_children():
        tree.delete(item)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Check if the ID matches the search term, if provided
        id_match = str(row[0]).zfill(3) == id_search.zfill(3) if id_search else True
        # Check if the Name matches the search term, if provided and not None
        name_match = name_search.lower() in (row[1] or "").lower() if name_search else True
        # If both ID and Name match (or are not provided), insert the row into the Treeview
        if id_match and name_match:
            tree.insert("", tk.END, values=row)
# Function to refresh the Treeview with updated data
def refresh_data():
    for item in tree.get_children():
        tree.delete(item)
    load_data()
# Load the workbook and select the active sheet
wb = load_workbook('.\data.xlsx')
sheet = wb.active
root = tk.Tk()
root.geometry("1246x650")
style = ttk.Style(root)
root.tk.call("source", ".\\forest-dark.tcl")
style.theme_use("forest-dark")
root.title("Wedding Guest Book")
# Create the entry widgets
entry_id = ttk.Entry(root, width=40, font=('Battambang', 14, 'bold'))
entry_id.grid(row=0, column=2)
entry_name = ttk.Entry(root, width=40, font=('Battambang', 14, 'bold'))
entry_name.grid(row=1, column=2)
entry_usd = ttk.Entry(root, width=40, font=('Battambang', 14, 'bold'))
entry_usd.grid(row=2, column=2)
entry_riel = ttk.Entry(root, width=40, font=('Battambang', 14, 'bold'))
entry_riel.grid(row=3, column=2)
# Create labels to display the previous update data
label_prev_id = ttk.Label(root, text="Previous ID:  ", width=40,anchor="w", justify="left", font=('Battambang', 12, 'bold'))
label_prev_id.grid(row=0, column=3)
label_prev_name = ttk.Label(root, text="Previous Name:   ", width=40,anchor="w", justify="left", font=('Battambang', 12, 'bold'))
label_prev_name.grid(row=1, column=3)
label_prev_usd = ttk.Label(root, text="Previous USD:   ", width=40,anchor="w", justify="left", font=('Battambang', 12, 'bold'))
label_prev_usd.grid(row=2, column=3)
label_prev_riel = ttk.Label(root, text="Previous Riel:   ", width=40,anchor="w", justify="left", font=('Battambang', 12, 'bold'))
label_prev_riel.grid(row=3, column=3)
# Create the labels for totals
label_total_guests = ttk.Label(root, text="Total Guests: 0", font=('Battambang', 12, 'bold'))
label_total_guests.grid(row=8, column=1)
label_total_usd = ttk.Label(root, text="Total USD: 0", font=('Battambang', 12, 'bold'))
label_total_usd.grid(row=8, column=2)
label_total_riel = ttk.Label(root, text="Total Riel: 0", font=('Battambang', 12, 'bold'))
label_total_riel.grid(row=8, column=3)
# Create the label widgets
label_id = ttk.Label(root, text="ID", width=5,anchor="w", justify="left",font=('Battambang', 12, 'bold'))
label_id.grid(row=0, column=1)
label_name = ttk.Label(root, text="Name", width=5,anchor="w", justify="left",font=('Battambang', 12, 'bold'))
label_name.grid(row=1, column=1)
label_usd = ttk.Label(root, text="USD",anchor="w", width=5, justify="left",font=('Battambang', 12, 'bold'))
label_usd.grid(row=2, column=1)
label_riel = ttk.Label(root, text="Riel",anchor="w", width=5, justify="left",font=('Battambang', 12, 'bold'))
label_riel.grid(row=3, column=1)
style.configure('Custom.TButton', font=('Helvetica', 12, 'bold'))
#Button
button_search = ttk.Button(root, text="Search", command=search_data, width=20, style='Custom.TButton')
button_search.grid(row=5, column=1, padx=5, pady=5)
button_clear = ttk.Button(root, text="Clear", command=clear_entries, width=20, style='Custom.TButton')
button_clear.grid(row=5, column=3, padx=5, pady=5)
button_update = ttk.Button(root, text="Update", command=update_data, width=20, style='Custom.TButton')
button_update.grid(row=5, column=2, padx=5, pady=5)
open_excel_button = ttk.Button(root, text="Open Excel", command=open_excel, width=20, style='Custom.TButton')
open_excel_button.grid(row=5, column=4, padx=5, pady=5)

# Get the list of printers
printer_list = get_printers()

# Create a dropdown menu (Combobox) to list all printer names
printer_dropdown = ttk.Combobox(root, values=printer_list)
printer_dropdown.grid(row=8, column=4, padx=5, pady=5)

# Print button
print_button = tk.Button(root, text="Print Excel File", command=lambda: print_excel_file(printer_dropdown.get()))
print_button.grid(row=9, column=4, padx=5, pady=5)

# Create a Style object
style = ttk.Style(root)
# Configure the Treeview style
style.configure("Treeview",
                font=('Battambang', 12),  # Font size for data
                rowheight=25)  # Row height
# Configure the Treeview Heading style
style.configure("Treeview.Heading",
                font=('Battambang', 12))  # Font size
# Create the Treeview widget
tree = ttk.Treeview(root, columns=("ID", "Name", "USD", "Riel"), show='headings')
tree.heading("ID", text="ID")
tree.heading("Name", text="Name")
tree.heading("USD", text="USD")
tree.heading("Riel", text="Riel")
# Configure the columns' width and alignment
tree.column("ID", width=10, anchor='center')
tree.column("Name", width=360, anchor='w')
tree.column("USD", width=30, anchor='e')
tree.column("Riel", width=40, anchor='e')
tree.grid(row=7,column=1, columnspan=4, sticky='nsew')
# Bind the select event of the tree to the on_tree_select function
tree.bind('<<TreeviewSelect>>', on_tree_select)
# Load data into the Treeview
load_data()
# Run the main loop
root.mainloop() #ok

